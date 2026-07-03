"""
tools_excel.py — Tools that read your real OneDrive files in Waseem Data.
"""

import os
import glob
from datetime import datetime, date
import openpyxl

from agent_core import tool

WASEEM_DATA = os.environ.get(
    "WASEEM_DATA_PATH",
    "/Users/waseemali/Library/CloudStorage/OneDrive-AbdullahTurkeyAlduhayansonsforconstruction/Waseem Data"
)
PETTY_CASH  = os.path.join(WASEEM_DATA, "Task/Petty Cash Sheet.xlsx")
TIME_DEP    = os.path.join(WASEEM_DATA, "2025 time deposit & murabaha.xlsx")


def _norm(s) -> str:
    return str(s or "").strip().lower().replace("_", " ")


# ---------------------------------------------------------------------------
# Tool 1 — list every Excel file the agent can see
# ---------------------------------------------------------------------------

@tool(
    name="list_files",
    description="List all Excel files in Waseem Data so the agent knows what's available.",
    schema={"type": "object", "properties": {}, "required": []},
)
def list_files():
    pattern = os.path.join(WASEEM_DATA, "**", "*.xlsx")
    files = glob.glob(pattern, recursive=True)
    return [os.path.relpath(f, WASEEM_DATA) for f in sorted(files)]


# ---------------------------------------------------------------------------
# Tool 2 — read any Excel file by relative path
# ---------------------------------------------------------------------------

@tool(
    name="read_excel",
    description=(
        "Read rows from any Excel file in Waseem Data. "
        "Pass the relative path (from list_files) and optionally the sheet name. "
        "Returns up to 200 rows as a list of dicts keyed by column header."
    ),
    schema={
        "type": "object",
        "properties": {
            "path":       {"type": "string", "description": "Relative path from list_files, e.g. 'Task/Petty Cash Sheet.xlsx'"},
            "sheet":      {"type": "string", "description": "Sheet name (optional — defaults to first sheet)"},
            "max_rows":   {"type": "integer", "description": "Max rows to return (default 100, max 200)"},
            "keyword":    {"type": "string", "description": "Only return rows that contain this keyword anywhere"},
        },
        "required": ["path"],
    },
)
def read_excel(path: str, sheet: str = None, max_rows: int = 100, keyword: str = None):
    full = os.path.join(WASEEM_DATA, path)
    if not os.path.exists(full):
        return f"ERROR: file not found — {path}"

    wb = openpyxl.load_workbook(full, read_only=True, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active

    rows_out = []
    headers = None
    count = 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if all(v in (None, "") for v in row):
            continue
        if headers is None:
            headers = [str(v or f"col{j}").strip() for j, v in enumerate(row)]
            continue
        record = {headers[j]: v for j, v in enumerate(row) if j < len(headers)}
        if keyword and keyword.lower() not in " ".join(str(v) for v in record.values()).lower():
            continue
        rows_out.append(record)
        count += 1
        if count >= min(max_rows, 200):
            break

    wb.close()
    return rows_out if rows_out else f"No data rows found in {path} / {ws.title}"


# ---------------------------------------------------------------------------
# Tool 3 — list sheets inside a file
# ---------------------------------------------------------------------------

@tool(
    name="list_sheets",
    description="List all sheet names in an Excel file.",
    schema={
        "type": "object",
        "properties": {
            "path": {"type": "string", "description": "Relative path from list_files"},
        },
        "required": ["path"],
    },
)
def list_sheets(path: str):
    full = os.path.join(WASEEM_DATA, path)
    if not os.path.exists(full):
        return f"ERROR: file not found — {path}"
    wb = openpyxl.load_workbook(full, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names


# ---------------------------------------------------------------------------
# Tool 4 — search across ALL files for a keyword
# ---------------------------------------------------------------------------

@tool(
    name="search_files",
    description=(
        "Search across all Excel files in Waseem Data for a keyword. "
        "Returns matching rows with their source file and sheet."
    ),
    schema={
        "type": "object",
        "properties": {
            "keyword": {"type": "string", "description": "Word or phrase to search for"},
            "max_hits": {"type": "integer", "description": "Max results to return (default 50)"},
        },
        "required": ["keyword"],
    },
)
def search_files(keyword: str, max_hits: int = 50):
    pattern = os.path.join(WASEEM_DATA, "**", "*.xlsx")
    files = glob.glob(pattern, recursive=True)
    hits = []

    for fpath in sorted(files):
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            for shname in wb.sheetnames:
                ws = wb[shname]
                headers = None
                for row in ws.iter_rows(values_only=True):
                    if all(v in (None, "") for v in row):
                        continue
                    if headers is None:
                        headers = [str(v or f"col{j}").strip() for j, v in enumerate(row)]
                        continue
                    row_str = " ".join(str(v) for v in row if v is not None)
                    if keyword.lower() in row_str.lower():
                        record = {headers[j]: v for j, v in enumerate(row) if j < len(headers)}
                        record["_file"] = os.path.relpath(fpath, WASEEM_DATA)
                        record["_sheet"] = shname
                        hits.append(record)
                        if len(hits) >= max_hits:
                            wb.close()
                            return hits
            wb.close()
        except Exception as e:
            hits.append({"_file": os.path.relpath(fpath, WASEEM_DATA), "_error": str(e)})

    return hits if hits else f"No results found for '{keyword}'"


# ---------------------------------------------------------------------------
# Tool 5 — petty cash summary (most-used file)
# ---------------------------------------------------------------------------

@tool(
    name="petty_cash_summary",
    description="Get recent petty cash transactions and current balance from Petty Cash Sheet.xlsx.",
    schema={
        "type": "object",
        "properties": {
            "last_n_rows": {"type": "integer", "description": "How many recent rows to return (default 20)"},
        },
        "required": [],
    },
)
def petty_cash_summary(last_n_rows: int = 20):
    if not os.path.exists(PETTY_CASH):
        return "ERROR: Petty Cash Sheet.xlsx not found"

    wb = openpyxl.load_workbook(PETTY_CASH, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    header = None

    for row in ws.iter_rows(values_only=True):
        if all(v in (None, "") for v in row):
            continue
        if header is None and any(v is not None for v in row[:6]):
            if "Date" in str(row[1]) or "date" in str(row[1]).lower() or "Decription" in str(row[2]):
                header = [str(v or f"col{i}") for i, v in enumerate(row)]
                continue
        if header:
            rows.append(row)

    wb.close()
    recent = rows[-last_n_rows:]
    if not header:
        return "Could not parse header row in Petty Cash Sheet"

    out = []
    for r in recent:
        record = {header[i]: v for i, v in enumerate(r) if i < len(header) and v not in (None, "")}
        out.append(record)

    balance = None
    for r in reversed(rows):
        if r[4] not in (None, ""):
            balance = r[4]
            break

    return {"balance": balance, "recent_transactions": out}
